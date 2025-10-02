# # from telegram import Update
# # from telegram.ext import Application, CommandHandler, ContextTypes

# # TOKEN = "8470587261:AAFSLT4uWXd9iuC-r5wv1XwEHvv8L4qI-AQ"  # token bot của bạn

# # # /start
# # async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
# #     await update.message.reply_text("Nhập theo cú pháp: /nhap <TênNgườiDùng> <TàiKhoản> <SốNguyên>")

# # # /nhap
# # async def nhap(update: Update, context: ContextTypes.DEFAULT_TYPE):
# #     if len(context.args) < 3:
# #         await update.message.reply_text("Sai cú pháp! /nhap <TÊN NGƯỜI DÙNG> <TÊN TK> <MỆNH GIÁ TIỀN>")
# #         return
    
# #     # tên người dùng luôn in hoa
# #     ten_nguoi_dung = context.args[0].upper()
# #     tai_khoan = context.args[1]
# #     try:
# #         gia_tri = int(context.args[2])
# #     except ValueError:
# #         await update.message.reply_text("Giá trị phải là số nguyên!")
# #         return

# #     ket_qua = f"/W {tai_khoan} - OWS {ten_nguoi_dung} - {gia_tri} - 5D"
# #     await update.message.reply_text(ket_qua)

# # def main():
# #     app = Application.builder().token(TOKEN).build()
# #     app.add_handler(CommandHandler("start", start))
# #     app.add_handler(CommandHandler("nhap", nhap))
# #     app.run_polling()

# # if __name__ == "__main__":
# #     main()import sqlite3


import os
from datetime import datetime
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes
from openpyxl import Workbook, load_workbook

TOKEN = "8470587261:AAFSLT4uWXd9iuC-r5wv1XwEHvv8L4qI-AQ"
FILE_NAME = "ketqua.xlsx"

# Hàm lưu vào Excel
def save_to_excel(nguoi_dung, tai_khoan, gia_tri, ket_qua):
    thoi_gian = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Nếu file chưa tồn tại -> tạo mới
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Nhập dữ liệu"
        ws.append(["Thời gian", "Người dùng", "Tài khoản", "Giá trị", "Kết quả"])
        wb.save(FILE_NAME)

    # Ghi thêm dữ liệu mới
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([thoi_gian, nguoi_dung, tai_khoan, gia_tri, ket_qua])
    wb.save(FILE_NAME)

# /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Chào bạn! Nhập theo cú pháp:\n"
        "/nhap <TênNgườiDùng> <TàiKhoản> <SốNguyên>\n"
        "Hoặc xem kết quả theo tên:\n"
        "/xem <TênNgườiDùng>"
    )

# /nhap
async def nhap(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 3:
        await update.message.reply_text("❌ Sai cú pháp!\nVí dụ: /nhap Kunz hiha123 123")
        return
    
    ten_nguoi_dung = context.args[0].upper()
    tai_khoan = context.args[1]

    try:
        gia_tri = int(context.args[2])
    except ValueError:
        await update.message.reply_text("❌ Giá trị phải là số nguyên!")
        return

    ket_qua = f"/W {tai_khoan} - OWS {ten_nguoi_dung} - {gia_tri} - 5D"

    # Lưu vào Excel
    save_to_excel(ten_nguoi_dung, tai_khoan, gia_tri, ket_qua)

    # Gửi 1 tin nhắn duy nhất
    await update.message.reply_text(f"\n{ket_qua}")
# /xem <TênNgườiDùng>
async def xem(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 1:
        await update.message.reply_text("❌ Vui lòng nhập tên người dùng để tìm kiếm.\nVí dụ: /xem Kunz")
        return
    
    ten_tim = context.args[0].upper()

    if not os.path.exists(FILE_NAME):
        await update.message.reply_text("❌ Chưa có dữ liệu nào trong ketqua.xlsx.")
        return

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ket_qua_tim = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, nguoi_dung, _, _, ket_qua = row
        if nguoi_dung.upper() == ten_tim:
            ket_qua_tim.append(ket_qua)

    if ket_qua_tim:
        # Gửi từng ket_qua 1 tin nhắn riêng
        for kq in ket_qua_tim:
            await update.message.reply_text(kq)
    else:
        await update.message.reply_text(f"❌ Không tìm thấy kết quả cho '{ten_tim}'.")
# /xoa <TênNgườiDùng>
async def xoa(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 1:
        await update.message.reply_text("❌ Vui lòng nhập tên người dùng để xóa.\nVí dụ: /xoa Kunz")
        return
    
    ten_xoa = context.args[0].upper()

    if not os.path.exists(FILE_NAME):
        await update.message.reply_text("❌ Chưa có dữ liệu nào trong ketqua.xlsx.")
        return

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    # Lọc lại các dòng không bị xóa
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows_giu_lai = [row for row in rows if row[1].upper() != ten_xoa]  # cột 1 là "Người dùng"

    so_dong_xoa = len(rows) - len(rows_giu_lai)

    if so_dong_xoa == 0:
        await update.message.reply_text(f"❌ Không tìm thấy dữ liệu của '{ten_xoa}' để xóa.")
        return

    # Xóa tất cả dòng cũ
    ws.delete_rows(2, ws.max_row - 1)

    # Ghi lại các dòng còn giữ lại
    for row in rows_giu_lai:
        ws.append(row)
    wb.save(FILE_NAME)

    await update.message.reply_text(f"✅ Đã xóa {so_dong_xoa} kết quả của '{ten_xoa}' thành công!")


# Main
def main():
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("nhap", nhap))
    app.add_handler(CommandHandler("xem", xem))
    app.run_polling()

if __name__ == "__main__":
    main()
